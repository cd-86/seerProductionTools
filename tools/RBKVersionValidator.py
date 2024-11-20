import json
import os.path
import socket

from PySide6.QtCore import QDir, QRegularExpression, QThread, Qt
from PySide6.QtGui import QRegularExpressionValidator, QStandardItemModel, QStandardItem, QFont, QPainter, QPainterPath, \
    QColor, QPen
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTableView, \
    QGridLayout, QSpacerItem, QFrame
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from init import cfg, printLog
from lib.RBKUtils import RBKUtils


class Thread(QThread):
    def __init__(self):
        super().__init__()
        self.ip = "192.168.192.5"
        self.rbkVersion = ""
        self.dspVersion = ""
        self.gyroVersion = ""
        self.srcPatch = ""
        self.robodVersion = ""
        self.srcName = ""
        self.isSuccess = False

    def run(self):
        self.isSuccess = False
        try:
            printLog(f"连接 {self.ip}:19204")
            so: socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            so.connect((self.ip, 19204))
            so.settimeout(60)
            printLog("查询机器人信息")
            data = RBKUtils.request(so, 1000)
            robot_status_info = json.loads(data)
            self.rbkVersion = robot_status_info.get('version', '')
            self.dspVersion = robot_status_info.get('dsp_version', "")
            self.gyroVersion = robot_status_info.get('gyro_version', "")
            for k, v in robot_status_info.get('VERSION_LIST', {}).items():
                if k == "x86-patch" or k == "arm-patch":
                    self.srcPatch = v
                    break
            else:
                self.srcPatch = ""
        except Exception as e:
            printLog("Exception:,", e)
            return
        else:
            printLog(f"关闭连接 {self.ip}:19204")
            so.close()
        try:
            printLog(f"连接 {self.ip}:19208")
            so: socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            so.connect((self.ip, 19208))
            so.settimeout(60)
            printLog("查询 Robod 版本")
            data = RBKUtils.request(so, 5041)
            robot_core_robod_version_info = json.loads(data)

            self.srcName = robot_core_robod_version_info.get("srcName")
            if not self.srcName:
                srcType = robot_core_robod_version_info.get("SRCType", -1)
                if 0 <= srcType <= 4:
                    self.srcName = ["[SRC-2000]", "[SRC-3000]", "[SRC-800]", "[SRC-2000.2]", "[SRC-2000.1]"][srcType]
                else:
                    self.srcName = "[SRC-?]"
            self.robodVersion = robot_core_robod_version_info.get('version', '')

        except Exception as e:
            printLog("Exception:,", e)
            return
        else:
            printLog(f"关闭连接 {self.ip}:19208")
            so.close()
        printLog("SRC:%s,RBK:%s,DPS:%s,GYRO:%s,SRC-PATCH:%s,ROBOD:%s"%(self.srcName, self.rbkVersion, self.dspVersion, self.gyroVersion, self.srcPatch, self.robodVersion))
        self.isSuccess = True


class ResultIcon(QWidget):
    def __init__(self):
        super().__init__()
        self.value: bool = None

    def paintEvent(self, event):
        if self.value is None:
            return
        w = min(self.width(), self.height())
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.translate((self.width() - w) / 2, (self.height() - w) / 2)
        if self.value:
            path = QPainterPath()
            path.moveTo(w * 0.1, w * 0.5)
            path.lineTo(w * 0.4, w * 0.8)
            path.lineTo(w * 0.9, w * 0.2)
            painter.setPen(QPen(QColor(25, 200, 25, 255), 4, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap,
                                Qt.PenJoinStyle.RoundJoin))
            painter.drawPath(path)
        else:
            painter.setPen(QPen(QColor(200, 25, 25, 255), 4, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap,
                                Qt.PenJoinStyle.RoundJoin))
            painter.drawLine(w * 0.2, w * 0.2, w * 0.8, w * 0.8)
            painter.drawLine(w * 0.2, w * 0.8, w * 0.8, w * 0.2)


class RBKVersionValidator(QWidget):
    META = {
        "title": "RBK 版本验证",
        "config": {
            "rbk_version_excel": {
                "name": "RBK 版本配置表路径",
                "desc": "Rbk 版本配置信息",
                "value": QDir.current().absolutePath() + "/RBKVersion/version.xlsx",
                "filter": "xlsx file (*.xlsx)",
                "type": "file"
            }
        }
    }

    def __init__(self):
        super().__init__()
        self.initUI()

        self.workThread = Thread()
        self.workThread.finished.connect(self.slotWorkThreadFinished)
        self.srcTypeList = []

    def initUI(self):
        layout = QVBoxLayout(self)

        hLayout = QHBoxLayout()
        hLayout.addWidget(QLabel("IP:"))
        self.ipLineEdit = QLineEdit("192.168.192.5")
        self.ipLineEdit.setValidator(QRegularExpressionValidator(
            QRegularExpression(r"((2(5[0-5]|[0-4]\d))|[0-1]?\d{1,2})(\.((2(5[0-5]|[0-4]\d))|[0-1]?\d{1,2})){3}"), self))
        hLayout.addWidget(self.ipLineEdit)
        self.loadVersionSheetButton = QPushButton("重新加载版本配置表")
        self.validateButton = QPushButton("验证版本信息")
        hLayout.addWidget(self.loadVersionSheetButton)
        hLayout.addWidget(self.validateButton)
        layout.addLayout(hLayout)

        self.tableView = QTableView()
        self.tableView.setEditTriggers(QTableView.EditTrigger.NoEditTriggers)
        self.tableView.horizontalHeader().setStretchLastSection(True)
        self.tableView.verticalHeader().setHidden(True)
        self.model = QStandardItemModel()
        self.tableView.setModel(self.model)
        layout.addWidget(self.tableView)
        layout.addItem(QSpacerItem(0, 100))
        self.resultLabel = QLabel("验证结果：")
        font = self.resultLabel.font()
        font.setPixelSize(20)
        self.resultLabel.setFont(font)
        layout.addWidget(self.resultLabel)

        self.gridLayout = QGridLayout()
        self.gridLayout.setHorizontalSpacing(40)
        self.gridLayout.setAlignment(Qt.AlignmentFlag.AlignLeft)

        def fn(s, align=Qt.AlignmentFlag.AlignRight):
            l = QLabel(s)
            l.setObjectName("Cell")
            l.setAlignment(align)
            font = l.font()
            font.setBold(True)
            l.setFont(font)
            return l

        self.srcTitle = fn("SRC", Qt.AlignmentFlag.AlignCenter)
        self.gridLayout.addWidget(self.srcTitle, 0, 1)
        self.gridLayout.addWidget(fn("值", Qt.AlignmentFlag.AlignCenter), 0, 2)
        self.gridLayout.addWidget(fn("是否一致", Qt.AlignmentFlag.AlignCenter), 0, 3)


        self.gridLayout.addWidget(fn("Robokit:"), 1, 0)
        self.gridLayout.addWidget(fn("Robod:"), 2, 0)
        self.gridLayout.addWidget(fn("SRC-PATCH:"), 3, 0)
        self.gridLayout.addWidget(fn("底层固件:"), 4, 0)
        self.gridLayout.addWidget(fn("陀螺仪固件:"), 5, 0)

        self.l11 = QLabel()
        self.l21 = QLabel()
        self.l31 = QLabel()
        self.l41 = QLabel()
        self.l51 = QLabel()
        self.gridLayout.addWidget(self.l11, 1, 1)
        self.gridLayout.addWidget(self.l21, 2, 1)
        self.gridLayout.addWidget(self.l31, 3, 1)
        self.gridLayout.addWidget(self.l41, 4, 1)
        self.gridLayout.addWidget(self.l51, 5, 1)

        self.l12 = QLabel()
        self.l22 = QLabel()
        self.l32 = QLabel()
        self.l42 = QLabel()
        self.l52 = QLabel()
        self.gridLayout.addWidget(self.l12, 1, 2)
        self.gridLayout.addWidget(self.l22, 2, 2)
        self.gridLayout.addWidget(self.l32, 3, 2)
        self.gridLayout.addWidget(self.l42, 4, 2)
        self.gridLayout.addWidget(self.l52, 5, 2)


        self.r13 = ResultIcon()
        self.r23 = ResultIcon()
        self.r33 = ResultIcon()
        self.r43 = ResultIcon()
        self.r53 = ResultIcon()
        self.gridLayout.addWidget(self.r13, 1, 3)
        self.gridLayout.addWidget(self.r23, 2, 3)
        self.gridLayout.addWidget(self.r33, 3, 3)
        self.gridLayout.addWidget(self.r43, 4, 3)
        self.gridLayout.addWidget(self.r53, 5, 3)

        layout.addLayout(self.gridLayout)

        self.loadVersionSheetButton.clicked.connect(self.slotLoadVersionSheetButtonClicked)
        self.validateButton.clicked.connect(self.slotValidateButtonClicked)
        self.loadVersionSheetButton.click()


    def slotLoadVersionSheetButtonClicked(self):
        self.model.clear()
        xlsx = cfg[self.__class__.__name__, "rbk_version_excel"]
        if not xlsx:
            xlsx = RBKVersionValidator.META["config"]["rbk_version_excel"]["value"]
        if not os.path.exists(xlsx):
            printLog(f"没有找到版本配置表 {xlsx}")
            return
        try:
            wb = load_workbook(xlsx)
            ws: Worksheet = wb.active
            self.srcTypeList = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
            self.model.setHorizontalHeaderLabels(self.srcTypeList)
            for row in range(2, 7):
                self.model.appendRow([QStandardItem(ws.cell(row, column).value) for column in range(1, ws.max_column + 1)])
            width = 1.0 / 5.0 * self.tableView.width()
            self.tableView.setColumnWidth(0, width)
            self.tableView.setColumnWidth(1, width)
            self.tableView.setColumnWidth(2, width)
            self.tableView.setColumnWidth(3, width)
            self.tableView.setColumnWidth(4, width)
            wb.close()
        except:
            printLog(f"读取版本配置表失败 {xlsx}")
            return
        printLog(f"加载版本配置表成功！")

    def slotValidateButtonClicked(self):
        self.workThread.ip = self.ipLineEdit.text()
        self.workThread.start()
        self.validateButton.setEnabled(False)

    def slotWorkThreadFinished(self):
        self.validateButton.setEnabled(True)
        if not self.workThread.isSuccess:
            return

        self.l12.setText(self.workThread.rbkVersion)
        self.l22.setText(self.workThread.robodVersion)
        self.l32.setText(self.workThread.srcPatch)
        self.l42.setText(self.workThread.dspVersion)
        self.l52.setText(self.workThread.gyroVersion)

        self.srcTitle.setText(self.workThread.srcName)
        for i, t in enumerate(self.srcTypeList):
            if t and t.upper() == self.workThread.srcName.upper():
                self.l11.setText(self.model.index(0, i).data(Qt.ItemDataRole.DisplayRole))
                self.l21.setText(self.model.index(1, i).data(Qt.ItemDataRole.DisplayRole))
                self.l31.setText(self.model.index(2, i).data(Qt.ItemDataRole.DisplayRole))
                self.l41.setText(self.model.index(3, i).data(Qt.ItemDataRole.DisplayRole))
                self.l51.setText(self.model.index(4, i).data(Qt.ItemDataRole.DisplayRole))
                break
        else:
            printLog(self.workThread.srcName, "Error")
            self.l11.setText("")
            self.l21.setText("")
            self.l31.setText("")
            self.l41.setText("")
            self.l51.setText("")

        self.r13.value = self.l11.text() == self.l12.text() if self.l11.text() and self.l12.text() else False
        self.r23.value = self.l21.text() == self.l22.text() if self.l21.text() and self.l22.text() else False
        self.r33.value = self.l31.text() == self.l32.text() if self.l31.text() and self.l32.text() else False
        self.r43.value = self.l41.text() == self.l42.text() if self.l41.text() and self.l42.text() else False
        self.r53.value = self.l51.text() == self.l52.text() if self.l51.text() and self.l52.text() else False

        result = self.r13.value and self.r23.value and self.r33.value and self.r43.value and self.r53.value
        printLog(f"验证结果:{'OK' if result else 'NG'}")
        s = '<b style="color: green;">OK</b>' if result else '<b style="color: red;">NG</b>'
        self.resultLabel.setText(f'<b>验证结果：</b>{s}')
